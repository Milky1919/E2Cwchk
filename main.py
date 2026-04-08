import os
import re
import csv
import math
from io import BytesIO, StringIO
from fastapi import FastAPI, UploadFile, File, Request, Form
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
import openpyxl

app = FastAPI(title="Excel Data Extractor")
templates = Jinja2Templates(directory="templates")

EXPECTED_CELLS = {
    "C7": "利用時間",
    "I7": "就労時間",
    "C8": "開始",
    "F8": "終了",
    "I8": "開始",
    "L8": "終了"
}

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    return templates.TemplateResponse(request=request, name="index.html", context={"error": None})

@app.post("/upload")
async def upload_file(request: Request, file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        return templates.TemplateResponse(request=request, name="index.html", context={"error": "Excelファイル(.xlsx, .xls)をアップロードしてください。"})
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as e:
        return templates.TemplateResponse(request=request, name="index.html", context={"error": f"ファイルの読み込みに失敗しました: {str(e)}"})

    if not wb.sheetnames:
        return templates.TemplateResponse(request=request, name="index.html", context={"error": "シートが見つかりません。"})

    sheet1_A1 = wb.worksheets[0]["A1"].value
    if sheet1_A1 is None:
        sheet1_A1 = ""

    sheet_results = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # バリデーション
        for cell_ref, expected in EXPECTED_CELLS.items():
            cell_val = sheet[cell_ref].value
            val = str(cell_val).replace('\n', '').strip() if cell_val is not None else ""
            if val != expected:
                error_msg = f"シート「{sheet_name}」の {cell_ref} セルが不正です。期待値: '{expected}', 実際の値: '{val}'"
                return templates.TemplateResponse(request=request, name="index.html", context={"error": error_msg})
        
        # シート名をそのまま名前に使用
        name = sheet_name
        
        diffs = []
        
        # 9行目から走査
        for r in range(9, sheet.max_row + 1):
            cell_a = sheet.cell(row=r, column=1).value
            if cell_a is None:
                val_a = ""
            else:
                val_a = str(cell_a).strip()
            
            # 「合計」が出たら終了
            if "合計" in val_a:
                break
                
            # 数字（日付）かチェック
            is_num = False
            if isinstance(cell_a, (int, float)) and not math.isnan(float(cell_a)):
                is_num = True
                date_val = str(int(cell_a))
            elif val_a.isdigit():
                is_num = True
                date_val = val_a
                
            if not is_num:
                # 28〜31以外の数字ではない文字が出たら終了
                break
                
            # 値の取得（空セル対応）
            def get_val(col_idx):
                v = sheet.cell(row=r, column=col_idx).value
                return str(v).strip() if v is not None else ""
                
            val_c = get_val(3)
            val_f = get_val(6)
            val_i = get_val(9)
            val_l = get_val(12)
            
            if val_c != val_i:
                diffs.append({"date": date_val, "type": "開始", "usage": val_c, "work": val_i})
            if val_f != val_l:
                diffs.append({"date": date_val, "type": "終了", "usage": val_f, "work": val_l})
                
        sheet_results.append({
            "name": name,
            "diffs": diffs
        })

    # CSV生成
    output = StringIO()
    writer = csv.writer(output)
    
    # 全シート共通: 1行目に1枚目のシートのA1
    writer.writerow([sheet1_A1])
    
    for res in sheet_results:
        # 差異があるシートのみ名前行とデータを書き込む
        if res["diffs"]:
            writer.writerow([res["name"]])
            for d in res["diffs"]:
                writer.writerow([d["date"], d["type"], d["usage"], d["work"]])
                
    output.seek(0)
    # cp932 エンコーディングの指定などが必要かどうか。要件に指定はないが、Excelで開くことを考慮してutf-8(with BOM)として返すのが無難。
    # PythonのStringIOでは文字列として処理し、StreamingResponseでUTF-8に変換時にBOMをつける。
    
    csv_bytes = output.getvalue().encode('utf-8-sig') # Excelで文字化けしないように BOM付き UTF-8
    
    response = StreamingResponse(BytesIO(csv_bytes), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=extracted_data.csv"
    return response

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=18923)
